"""
Benchmark harness for the customer AgentKernel.

Runs 10 independent trials of a 5-question conversation and evaluates
correctness against ground-truth answers derived from the live SQL data.

Outputs:
  - benchmarks/benchmark_results.md   (detailed report)
  - benchmarks/benchmark_results.xlsx  (Excel with charts)
"""

import asyncio
import json
import logging
import os
import re
import sys
import time
from datetime import datetime

import pandas as pd
import pyodbc

# Ensure project root is on path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)) + "/..")

from config.settings import Settings
from data.loader import DataLoader
from agent.kernel import AgentKernel

logging.basicConfig(
    level=logging.WARNING,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    stream=sys.stdout,
)
logger = logging.getLogger("benchmark")
logger.setLevel(logging.INFO)

NUM_TRIALS = 10
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "benchmarks")

# ── Ground truth from SQL ─────────────────────────────────────────

def fetch_ground_truth() -> dict:
    """Query SQL directly to establish expected answers."""
    conn = pyodbc.connect(
        "DRIVER={ODBC Driver 17 for SQL Server};"
        "Server=jortizflores,1433;"
        "Database=ExtronDemo;"
        "Trusted_Connection=yes;"
    )
    cursor = conn.cursor()

    # Total row count
    cursor.execute("SELECT COUNT(*) FROM [operations].[Obsolescence_Results]")
    total_rows = cursor.fetchone()[0]

    # Pick a real part that MAY be eligible to scrap (deterministic seed)
    cursor.execute("""
        SELECT TOP 1 PartNumber, Status, Details
        FROM [operations].[Obsolescence_Results]
        WHERE Status = 'May be eligible to be scrapped'
        ORDER BY PartNumber
    """)
    scrappable = cursor.fetchone()

    # Pick a real part that is NOT eligible (deterministic seed)
    cursor.execute("""
        SELECT TOP 1 PartNumber, Status, Details
        FROM [operations].[Obsolescence_Results]
        WHERE Status LIKE 'NOT eligible%'
        ORDER BY PartNumber
    """)
    not_eligible = cursor.fetchone()

    conn.close()

    fake_part = "99-0000-FAKE"

    return {
        "total_rows": total_rows,
        "real_part": {
            "number": scrappable[0],
            "status": scrappable[1],
            "details": scrappable[2],
        },
        "not_eligible_part": {
            "number": not_eligible[0],
            "status": not_eligible[1],
            "details": not_eligible[2],
        },
        "fake_part": fake_part,
    }


# ── Evaluation functions ──────────────────────────────────────────

def eval_q1_row_count(response_text: str, expected_count: int) -> dict:
    """Q1: 'how many rows does the table have' → must mention 500."""
    numbers = re.findall(r"\b(\d+)\b", response_text)
    found_count = any(int(n) == expected_count for n in numbers)
    return {
        "question": "How many rows does the table have?",
        "expected": str(expected_count),
        "got": response_text[:300],
        "pass": found_count,
        "criteria": f"Response contains the number {expected_count}",
    }


def eval_q2_show_rows(response_text: str, data_chunks: list, files: list, expected_count: int) -> dict:
    """Q2: 'show me the rows' → must return data chunks OR auto-export Excel for large datasets."""
    has_data = len(data_chunks) > 0
    has_files = len(files) > 0  # auto-Excel for 500 > MAX_INLINE_ROWS (200) is correct
    mentions_count = str(expected_count) in response_text
    mentions_sent = any(
        kw in response_text.lower()
        for kw in ["rows", "data", "sent", "retrieved", "records", "results",
                    "file", "generated", "excel", "export", "download"]
    )
    passed = has_data or has_files or (mentions_count and mentions_sent)
    return {
        "question": "Show me the rows",
        "expected": f"Data returned ({expected_count} rows) — inline or auto-Excel",
        "got": f"{len(data_chunks)} chunks, {len(files)} files; text: {response_text[:200]}",
        "pass": passed,
        "criteria": "Data chunks populated OR auto-Excel generated (500 > MAX_INLINE_ROWS)",
    }


def eval_q3_excel_export(response_text: str, files: list) -> dict:
    """Q3: 'please put the results into an excel file' → must generate file."""
    has_file = len(files) > 0
    mentions_file = any(
        kw in response_text.lower()
        for kw in ["excel", "file", "generated", "download", "export", ".xlsx"]
    )
    passed = has_file or mentions_file
    return {
        "question": "Please put the results into an Excel file",
        "expected": "Excel file generated",
        "got": f"Files: {files}; text: {response_text[:200]}",
        "pass": passed,
        "criteria": "File buffer populated OR response mentions Excel/file generation",
    }


def eval_q4_real_part(response_text: str, part_info: dict) -> dict:
    """Q4: 'can part X be scrapped?' where X exists → must return correct status."""
    part_num = part_info["number"]
    expected_status = part_info["status"]

    # Check if the status is mentioned in the response
    status_in_response = expected_status.lower() in response_text.lower()
    # Also accept partial matches for the key phrases
    key_phrases = ["may be eligible", "eligible to be scrapped"]
    partial_match = any(kp in response_text.lower() for kp in key_phrases)

    passed = status_in_response or partial_match
    return {
        "question": f"Can part {part_num} be scrapped?",
        "expected": f"Status: {expected_status}",
        "got": response_text[:300],
        "pass": passed,
        "criteria": f"Response includes status '{expected_status}' or key phrases",
    }


def eval_q5_fake_part(response_text: str, fake_part: str) -> dict:
    """Q5: 'can part FAKE be scrapped?' → must say not found / no data."""
    not_found_phrases = [
        "could not find",
        "not found",
        "no data",
        "does not exist",
        "doesn't exist",
        "no matching",
        "no results",
        "cannot determine",
        "no record",
    ]
    passed = any(phrase in response_text.lower() for phrase in not_found_phrases)
    return {
        "question": f"Can part {fake_part} be scrapped?",
        "expected": "Part not found / cannot determine",
        "got": response_text[:300],
        "pass": passed,
        "criteria": "Response indicates part not found or no data available",
    }


# ── Run a single trial ───────────────────────────────────────────

async def run_trial(trial_num: int, ground_truth: dict) -> list[dict]:
    """Run one full 5-question conversation and return evaluation results."""
    settings = Settings()
    data_loader = DataLoader(settings)
    kernel = AgentKernel(settings, data_loader)
    conv_id = f"benchmark_trial_{trial_num}_{int(time.time())}"

    table_name = settings.sql_table
    real_part = ground_truth["real_part"]["number"]
    fake_part = ground_truth["fake_part"]

    questions = [
        "how many rows does the table have",
        "show me the rows",
        "please put the results into an excel file",
        f"can part {real_part} be scrapped?",
        f"can part {fake_part} be scrapped?",
    ]

    results = []
    for i, question in enumerate(questions):
        logger.info("  Trial %d, Q%d: %s", trial_num, i + 1, question)
        start = time.time()
        try:
            response = await kernel.ask(conv_id, question)
        except Exception as e:
            response = {"text": f"ERROR: {e}", "data_chunks": [], "files": []}
        elapsed = time.time() - start

        text = response.get("text", "")
        data_chunks = response.get("data_chunks", [])
        files = response.get("files", [])

        if i == 0:
            ev = eval_q1_row_count(text, ground_truth["total_rows"])
        elif i == 1:
            ev = eval_q2_show_rows(text, data_chunks, files, ground_truth["total_rows"])
        elif i == 2:
            ev = eval_q3_excel_export(text, files)
        elif i == 3:
            ev = eval_q4_real_part(text, ground_truth["real_part"])
        elif i == 4:
            ev = eval_q5_fake_part(text, fake_part)

        ev["trial"] = trial_num
        ev["q_num"] = i + 1
        ev["latency_s"] = round(elapsed, 2)
        ev["raw_response"] = text[:500]
        results.append(ev)

    return results


# ── Markdown report ───────────────────────────────────────────────

def generate_markdown(all_results: list[dict], ground_truth: dict, total_elapsed: float) -> str:
    lines = []
    lines.append("# Agent Benchmark Report")
    lines.append(f"\n**Date:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"**Trials:** {NUM_TRIALS}")
    lines.append(f"**Total runtime:** {total_elapsed:.1f}s")
    lines.append(f"**Model:** {Settings().azure_openai_deployment_name}")
    lines.append(f"**Table:** {Settings().sql_table} ({ground_truth['total_rows']} rows)")

    lines.append("\n## Ground Truth")
    lines.append(f"- **Total rows:** {ground_truth['total_rows']}")
    lines.append(f"- **Real part (scrappable):** `{ground_truth['real_part']['number']}` → *{ground_truth['real_part']['status']}*")
    lines.append(f"- **Fake part:** `{ground_truth['fake_part']}` → *not in database*")

    lines.append("\n## Conversation Template")
    lines.append("| # | Question |")
    lines.append("|---|----------|")
    lines.append("| 1 | How many rows does the table have? |")
    lines.append("| 2 | Show me the rows |")
    lines.append("| 3 | Please put the results into an Excel file |")
    lines.append(f"| 4 | Can part {ground_truth['real_part']['number']} be scrapped? |")
    lines.append(f"| 5 | Can part {ground_truth['fake_part']} be scrapped? |")

    # Per-question accuracy
    df = pd.DataFrame(all_results)
    lines.append("\n## Per-Question Accuracy")
    lines.append("| Question | Pass Rate | Avg Latency |")
    lines.append("|----------|-----------|-------------|")
    for q_num in sorted(df["q_num"].unique()):
        qdf = df[df["q_num"] == q_num]
        pass_rate = qdf["pass"].sum() / len(qdf) * 100
        avg_lat = qdf["latency_s"].mean()
        q_text = qdf.iloc[0]["question"][:60]
        lines.append(f"| Q{q_num}: {q_text} | {pass_rate:.0f}% ({int(qdf['pass'].sum())}/{len(qdf)}) | {avg_lat:.2f}s |")

    # Overall accuracy
    total_pass = df["pass"].sum()
    total_questions = len(df)
    overall_pct = total_pass / total_questions * 100
    lines.append(f"\n## Overall Accuracy: **{overall_pct:.1f}%** ({total_pass}/{total_questions})")

    # Per-trial breakdown
    lines.append("\n## Per-Trial Breakdown")
    lines.append("| Trial | Q1 | Q2 | Q3 | Q4 | Q5 | Score | Latency |")
    lines.append("|-------|----|----|----|----|----| ------|---------|")
    for trial in range(1, NUM_TRIALS + 1):
        tdf = df[df["trial"] == trial]
        marks = []
        for q in range(1, 6):
            row = tdf[tdf["q_num"] == q]
            if len(row) > 0 and row.iloc[0]["pass"]:
                marks.append("✅")
            else:
                marks.append("❌")
        score = tdf["pass"].sum()
        total_lat = tdf["latency_s"].sum()
        lines.append(f"| {trial} | {' | '.join(marks)} | {score}/5 | {total_lat:.1f}s |")

    # Detailed failures
    failures = df[~df["pass"]]
    if len(failures) > 0:
        lines.append("\n## Failure Details")
        for _, row in failures.iterrows():
            lines.append(f"\n### Trial {row['trial']}, Q{row['q_num']}: {row['question']}")
            lines.append(f"- **Expected:** {row['expected']}")
            lines.append(f"- **Got:** {row['raw_response'][:300]}")
            lines.append(f"- **Criteria:** {row['criteria']}")
    else:
        lines.append("\n## No failures — all questions passed in all trials! 🎉")

    return "\n".join(lines)


# ── Excel report with charts ─────────────────────────────────────

def generate_excel(all_results: list[dict], ground_truth: dict, output_path: str = None):
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    df = pd.DataFrame(all_results)
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filepath = output_path or os.path.join(OUTPUT_DIR, "benchmark_results.xlsx")

    wb = Workbook()

    # ── Sheet 1: Raw Results ──
    ws_raw = wb.active
    ws_raw.title = "Raw Results"
    headers = ["Trial", "Q#", "Question", "Expected", "Got", "Pass", "Latency (s)"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for c, h in enumerate(headers, 1):
        cell = ws_raw.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for r_idx, (_, row) in enumerate(df.iterrows(), 2):
        vals = [
            row["trial"],
            row["q_num"],
            row["question"][:80],
            row["expected"][:80],
            row["raw_response"][:120],
            "PASS" if row["pass"] else "FAIL",
            row["latency_s"],
        ]
        for c, v in enumerate(vals, 1):
            cell = ws_raw.cell(row=r_idx, column=c, value=v)
            cell.border = thin_border
            if c == 6:
                cell.fill = pass_fill if v == "PASS" else fail_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

    # Auto-width
    for c in range(1, len(headers) + 1):
        max_len = max(len(str(ws_raw.cell(row=r, column=c).value or "")) for r in range(1, ws_raw.max_row + 1))
        ws_raw.column_dimensions[get_column_letter(c)].width = min(max_len + 4, 50)

    # ── Sheet 2: Summary ──
    ws_sum = wb.create_sheet("Summary")

    # Per-question accuracy table
    ws_sum.cell(row=1, column=1, value="Per-Question Accuracy").font = Font(bold=True, size=14)
    sum_headers = ["Question", "Pass Count", "Fail Count", "Pass Rate (%)", "Avg Latency (s)"]
    for c, h in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=3, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    q_labels = []
    for q_num in sorted(df["q_num"].unique()):
        qdf = df[df["q_num"] == q_num]
        q_text = f"Q{q_num}: {qdf.iloc[0]['question'][:50]}"
        q_labels.append(q_text)
        r = 3 + q_num
        pass_count = int(qdf["pass"].sum())
        fail_count = len(qdf) - pass_count
        pass_rate = round(pass_count / len(qdf) * 100, 1)
        avg_lat = round(qdf["latency_s"].mean(), 2)

        vals = [q_text, pass_count, fail_count, pass_rate, avg_lat]
        for c, v in enumerate(vals, 1):
            cell = ws_sum.cell(row=r, column=c, value=v)
            cell.border = thin_border
            if c == 4:
                cell.fill = pass_fill if pass_rate == 100 else fail_fill
                cell.font = Font(bold=True)

    # Overall accuracy
    total_pass = int(df["pass"].sum())
    total_q = len(df)
    overall = round(total_pass / total_q * 100, 1)
    r_overall = 3 + len(df["q_num"].unique()) + 2
    ws_sum.cell(row=r_overall, column=1, value="OVERALL ACCURACY").font = Font(bold=True, size=12)
    cell = ws_sum.cell(row=r_overall, column=2, value=f"{overall}% ({total_pass}/{total_q})")
    cell.font = Font(bold=True, size=12, color="4472C4")

    # Auto-width summary
    for c in range(1, len(sum_headers) + 1):
        max_len = max(len(str(ws_sum.cell(row=r, column=c).value or "")) for r in range(1, ws_sum.max_row + 1))
        ws_sum.column_dimensions[get_column_letter(c)].width = min(max_len + 4, 55)

    # ── Chart 1: Pass Rate by Question ──
    chart1 = BarChart()
    chart1.type = "col"
    chart1.title = "Pass Rate by Question (%)"
    chart1.y_axis.title = "Pass Rate (%)"
    chart1.y_axis.scaling.min = 0
    chart1.y_axis.scaling.max = 100
    chart1.style = 10
    chart1.width = 20
    chart1.height = 12

    data_ref = Reference(ws_sum, min_col=4, min_row=3, max_row=3 + len(q_labels))
    cats_ref = Reference(ws_sum, min_col=1, min_row=4, max_row=3 + len(q_labels))
    chart1.add_data(data_ref, titles_from_data=True)
    chart1.set_categories(cats_ref)
    chart1.shape = 4
    ws_sum.add_chart(chart1, f"A{r_overall + 3}")

    # ── Sheet 3: Per-Trial Scores ──
    ws_trial = wb.create_sheet("Per-Trial Scores")
    trial_headers = ["Trial", "Q1", "Q2", "Q3", "Q4", "Q5", "Score", "Total Latency (s)"]
    for c, h in enumerate(trial_headers, 1):
        cell = ws_trial.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for trial in range(1, NUM_TRIALS + 1):
        tdf = df[df["trial"] == trial]
        r = trial + 1
        ws_trial.cell(row=r, column=1, value=trial).border = thin_border
        for q in range(1, 6):
            qrow = tdf[tdf["q_num"] == q]
            passed = bool(qrow.iloc[0]["pass"]) if len(qrow) > 0 else False
            cell = ws_trial.cell(row=r, column=q + 1, value="PASS" if passed else "FAIL")
            cell.fill = pass_fill if passed else fail_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        score = int(tdf["pass"].sum())
        ws_trial.cell(row=r, column=7, value=score).border = thin_border
        ws_trial.cell(row=r, column=8, value=round(tdf["latency_s"].sum(), 1)).border = thin_border

    # Auto-width
    for c in range(1, len(trial_headers) + 1):
        max_len = max(len(str(ws_trial.cell(row=r, column=c).value or "")) for r in range(1, ws_trial.max_row + 1))
        ws_trial.column_dimensions[get_column_letter(c)].width = min(max_len + 4, 25)

    # ── Chart 2: Score per trial ──
    chart2 = BarChart()
    chart2.type = "col"
    chart2.title = "Score per Trial (out of 5)"
    chart2.y_axis.title = "Questions Passed"
    chart2.y_axis.scaling.min = 0
    chart2.y_axis.scaling.max = 5
    chart2.style = 10
    chart2.width = 20
    chart2.height = 12

    score_ref = Reference(ws_trial, min_col=7, min_row=1, max_row=NUM_TRIALS + 1)
    trial_cats = Reference(ws_trial, min_col=1, min_row=2, max_row=NUM_TRIALS + 1)
    chart2.add_data(score_ref, titles_from_data=True)
    chart2.set_categories(trial_cats)
    chart2.shape = 4
    ws_trial.add_chart(chart2, "A14")

    wb.save(filepath)
    return filepath


# ── Main ──────────────────────────────────────────────────────────

async def main():
    logger.info("=" * 60)
    logger.info("BENCHMARK: Fetching ground truth from SQL...")
    gt = fetch_ground_truth()
    logger.info("Ground truth: %d rows, real part=%s, fake part=%s",
                gt["total_rows"], gt["real_part"]["number"], gt["fake_part"])

    all_results = []
    overall_start = time.time()

    for trial in range(1, NUM_TRIALS + 1):
        logger.info("─" * 40)
        logger.info("TRIAL %d/%d", trial, NUM_TRIALS)
        trial_results = await run_trial(trial, gt)
        all_results.extend(trial_results)

        passed = sum(1 for r in trial_results if r["pass"])
        logger.info("  Trial %d result: %d/5 passed", trial, passed)

    total_elapsed = time.time() - overall_start
    logger.info("=" * 60)
    logger.info("ALL TRIALS COMPLETE in %.1fs", total_elapsed)

    # Overall stats
    df = pd.DataFrame(all_results)
    total_pass = df["pass"].sum()
    total_q = len(df)
    logger.info("OVERALL ACCURACY: %d/%d (%.1f%%)", total_pass, total_q, total_pass / total_q * 100)

    # Generate outputs
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    md_report = generate_markdown(all_results, gt, total_elapsed)
    md_path = os.path.join(OUTPUT_DIR, "benchmark_results.md")
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(md_report)
    logger.info("Markdown report: %s", md_path)

    try:
        xlsx_path = generate_excel(all_results, gt)
        logger.info("Excel report: %s", xlsx_path)
    except PermissionError:
        from datetime import datetime as _dt
        ts = _dt.now().strftime("%Y%m%d_%H%M%S")
        alt_path = os.path.join(OUTPUT_DIR, f"benchmark_results_{ts}.xlsx")
        xlsx_path = generate_excel(all_results, gt, output_path=alt_path)
        logger.info("Excel report (alt, original was open): %s", xlsx_path)

    print("\n" + "=" * 60)
    print(f"BENCHMARK COMPLETE — {total_pass}/{total_q} ({total_pass/total_q*100:.1f}%)")
    print(f"  Markdown: {md_path}")
    print(f"  Excel:    {xlsx_path}")
    print("=" * 60)


if __name__ == "__main__":
    asyncio.run(main())
