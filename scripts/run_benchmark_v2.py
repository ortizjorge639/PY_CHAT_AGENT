"""
Robust Agent Benchmark — v2

Tests the customer AgentKernel with:
  1. Varied phrasings of the same intent
  2. Multiple real parts (with & without Details, diverse statuses)
  3. Fake / non-existent part numbers
  4. Adversarial inputs (typos, near-misses, vague queries)
  5. Hallucination detection (asking for columns/statuses that don't exist)
  6. Multi-turn conversational coherence
  7. Strict evaluation criteria

Runs 10 independent trials. Each trial is a fresh kernel + session.
"""

import asyncio
import json
import logging
import math
import os
import random
import re
import sys
import time
from datetime import datetime
from typing import Any, Callable

import pandas as pd
import pyodbc

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))

from config.settings import Settings
from data.loader import DataLoader
from agent.kernel import AgentKernel

logging.basicConfig(
    level=logging.WARNING,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    stream=sys.stdout,
)
logger = logging.getLogger("benchmark_v2")
logger.setLevel(logging.INFO)

NUM_TRIALS = 10
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "benchmarks")

# ─────────────────────────────────────────────────────────────────
# Ground truth
# ─────────────────────────────────────────────────────────────────

def fetch_ground_truth() -> dict:
    conn = pyodbc.connect(
        "DRIVER={ODBC Driver 17 for SQL Server};"
        "Server=jortizflores,1433;"
        "Database=ExtronDemo;"
        "Trusted_Connection=yes;"
    )
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*) FROM [operations].[Obsolescence_Results]")
    total_rows = cursor.fetchone()[0]

    # One part per status (deterministic — ORDER BY PartNumber)
    cursor.execute("""
        SELECT PartNumber, Status, Details
        FROM (
            SELECT PartNumber, Status, Details,
                   ROW_NUMBER() OVER (PARTITION BY Status ORDER BY PartNumber) as rn
            FROM [operations].[Obsolescence_Results]
        ) sub WHERE rn = 1 ORDER BY Status
    """)
    parts_by_status = {}
    for row in cursor.fetchall():
        parts_by_status[row[1]] = {"number": row[0], "status": row[1], "details": row[2]}

    # Parts with non-null Details
    cursor.execute("""
        SELECT TOP 3 PartNumber, Status, Details
        FROM [operations].[Obsolescence_Results]
        WHERE Details IS NOT NULL
        ORDER BY PartNumber
    """)
    parts_with_details = [{"number": r[0], "status": r[1], "details": r[2]} for r in cursor.fetchall()]

    # Parts with null Details
    cursor.execute("""
        SELECT TOP 3 PartNumber, Status, Details
        FROM [operations].[Obsolescence_Results]
        WHERE Details IS NULL
        ORDER BY PartNumber
    """)
    parts_without_details = [{"number": r[0], "status": r[1], "details": r[2]} for r in cursor.fetchall()]

    # All part numbers for hallucination checks
    cursor.execute("SELECT PartNumber FROM [operations].[Obsolescence_Results]")
    all_parts = {r[0] for r in cursor.fetchall()}

    # Status distribution
    cursor.execute("""
        SELECT Status, COUNT(*) FROM [operations].[Obsolescence_Results]
        GROUP BY Status ORDER BY COUNT(*) DESC
    """)
    status_dist = {r[0]: r[1] for r in cursor.fetchall()}

    conn.close()
    return {
        "total_rows": total_rows,
        "parts_by_status": parts_by_status,
        "parts_with_details": parts_with_details,
        "parts_without_details": parts_without_details,
        "all_parts": all_parts,
        "status_distribution": status_dist,
    }


# ─────────────────────────────────────────────────────────────────
# Test question definitions
# ─────────────────────────────────────────────────────────────────

CATEGORIES = [
    "Row Count (varied phrasing)",
    "Part Lookup (with Details)",
    "Part Lookup (NULL Details)",
    "Part Lookup (NOT eligible status)",
    "Fake Part Rejection",
    "Status Distribution Query",
    "Excel Export Request",
    "Hallucination Resistance",
]


def build_test_questions(gt: dict, trial_num: int) -> list[dict]:
    """Build a conversation of ~8 questions with ground-truth + eval functions.

    Uses trial_num to rotate phrasings so each trial is different.
    """
    questions = []

    # ── Q1: Row count — varied phrasing ──────────────────────────
    count_phrasings = [
        "how many rows does the table have",
        "what's the total number of records?",
        "count the rows in the table",
        "how big is the dataset?",
        "what is the row count?",
        "tell me how many entries are in the table",
        "total number of parts in the table?",
        "how many parts are there?",
        "give me the record count",
        "what's the size of the data?",
    ]
    phrasing = count_phrasings[trial_num % len(count_phrasings)]
    questions.append({
        "category": "Row Count (varied phrasing)",
        "question": phrasing,
        "eval": lambda resp, gt=gt, **kw: _eval_row_count(resp, gt["total_rows"]),
    })

    # ── Q2: Part lookup WITH Details ─────────────────────────────
    parts_wd = gt["parts_with_details"]
    part_wd = parts_wd[trial_num % len(parts_wd)]
    lookup_phrasings = [
        f"can part {part_wd['number']} be scrapped?",
        f"what is the status of {part_wd['number']}?",
        f"look up part {part_wd['number']}",
        f"tell me about part number {part_wd['number']}",
        f"is {part_wd['number']} eligible for scrap?",
        f"check part {part_wd['number']} for me",
        f"what do we know about {part_wd['number']}?",
        f"pull up {part_wd['number']}",
        f"find part {part_wd['number']} in the database",
        f"scrap eligibility for {part_wd['number']}?",
    ]
    questions.append({
        "category": "Part Lookup (with Details)",
        "question": lookup_phrasings[trial_num % len(lookup_phrasings)],
        "eval": lambda resp, p=part_wd, **kw: _eval_part_lookup(resp, p, details_required=True),
    })

    # ── Q3: Part lookup WITHOUT Details (NULL) ───────────────────
    parts_nd = gt["parts_without_details"]
    part_nd = parts_nd[trial_num % len(parts_nd)]
    nd_phrasings = [
        f"can part {part_nd['number']} be scrapped?",
        f"what's the status of part {part_nd['number']}?",
        f"look up {part_nd['number']} for me",
        f"is part {part_nd['number']} scrap-eligible?",
        f"check {part_nd['number']}",
        f"what about {part_nd['number']}?",
        f"tell me the disposition of {part_nd['number']}",
        f"query part {part_nd['number']}",
        f"part {part_nd['number']} — can we scrap it?",
        f"status of {part_nd['number']} please",
    ]
    questions.append({
        "category": "Part Lookup (NULL Details)",
        "question": nd_phrasings[trial_num % len(nd_phrasings)],
        "eval": lambda resp, p=part_nd, **kw: _eval_part_lookup(resp, p, details_required=False),
    })

    # ── Q4: Part with NOT eligible status ────────────────────────
    not_eligible_statuses = [
        s for s in gt["parts_by_status"] if s.startswith("NOT eligible")
    ]
    ne_status = not_eligible_statuses[trial_num % len(not_eligible_statuses)]
    ne_part = gt["parts_by_status"][ne_status]
    questions.append({
        "category": "Part Lookup (NOT eligible status)",
        "question": f"can I scrap part {ne_part['number']}?",
        "eval": lambda resp, p=ne_part, **kw: _eval_not_eligible(resp, p),
    })

    # ── Q5: Fake part number ─────────────────────────────────────
    fake_parts = [
        "99-0000-FAKE",
        "00-0001-XX",
        "11-1111-11",
        "AA-BBBB-CC",
        "12-3456-NOPE",
        "99-9999-01LF",
        "00-0000-00",
        "55-0000-ZZ",
        "78-1234-BOGUS",
        "42-0000-TEST",
    ]
    fake = fake_parts[trial_num % len(fake_parts)]
    # Verify it's actually fake
    while fake in gt["all_parts"]:
        fake = f"99-{random.randint(1000,9999)}-FAKE"
    questions.append({
        "category": "Fake Part Rejection",
        "question": f"can part {fake} be scrapped?",
        "eval": lambda resp, f=fake, **kw: _eval_fake_part(resp, f),
    })

    # ── Q6: Status distribution / group-by ───────────────────────
    dist_phrasings = [
        "how many parts are in each status?",
        "break down the data by status",
        "give me a count per status",
        "show the status distribution",
        "group the parts by their status",
        "what's the breakdown by status?",
        "count parts grouped by status",
        "summarize by status",
        "how are parts distributed across statuses?",
        "status summary please",
    ]
    questions.append({
        "category": "Status Distribution Query",
        "question": dist_phrasings[trial_num % len(dist_phrasings)],
        "eval": lambda resp, gt=gt, **kw: _eval_status_distribution(resp, gt),
    })

    # ── Q7: Excel export (varied phrasing) ───────────────────────
    excel_phrasings = [
        "please put the results into an excel file",
        "export that to Excel",
        "can I get a spreadsheet download?",
        "generate an Excel file with those results",
        "download as xlsx please",
        "I need that in a spreadsheet",
        "create an Excel export",
        "give me a downloadable file",
        "save that as Excel",
        "export to spreadsheet",
    ]
    questions.append({
        "category": "Excel Export Request",
        "question": excel_phrasings[trial_num % len(excel_phrasings)],
        "eval": lambda resp, **kw: _eval_excel_export(resp),
    })

    # ── Q8: Hallucination resistance ─────────────────────────────
    hallucination_questions = [
        ("what is the price of part 15-3167-11?",
         "price", "There is no Price column"),
        ("show me the supplier for part 15-1539-01",
         "supplier", "There is no Supplier column"),
        ("how many parts have status 'Discontinued'?",
         "discontinued", "Discontinued is not a valid status"),
        ("what is the warranty status of part 15-1618-01A?",
         "warranty", "There is no warranty column"),
        ("list all parts from the New York warehouse",
         "warehouse", "There is no warehouse/location column beyond what exists"),
        ("which parts were manufactured in 2025?",
         "manufactured", "There is no manufacturing date column"),
        ("show me parts with priority level High",
         "priority", "There is no priority column"),
        ("what's the cost to scrap part 15-4578-02?",
         "cost", "There is no cost column"),
        ("who approved the scrap for part 15-3126-02LF?",
         "approved", "There is no approval column"),
        ("list parts from vendor Acme Corp",
         "vendor", "There is no vendor column"),
    ]
    hq = hallucination_questions[trial_num % len(hallucination_questions)]
    questions.append({
        "category": "Hallucination Resistance",
        "question": hq[0],
        "eval": lambda resp, keyword=hq[1], desc=hq[2], **kw: _eval_hallucination(resp, keyword, desc),
    })

    return questions


# ─────────────────────────────────────────────────────────────────
# Evaluation functions
# ─────────────────────────────────────────────────────────────────

def _eval_row_count(resp: dict, expected: int) -> dict:
    text = resp["text"]
    numbers = re.findall(r"\b(\d+)\b", text)
    found = any(int(n) == expected for n in numbers)
    return {
        "pass": found,
        "expected": str(expected),
        "got": text[:300],
        "criteria": f"Response contains exact number {expected}",
        "strict_notes": "Checks for exact count, not approximate.",
    }


def _eval_part_lookup(resp: dict, part_info: dict, details_required: bool) -> dict:
    text = resp["text"]
    part_num = part_info["number"]
    expected_status = part_info["status"]
    expected_details = part_info["details"]

    checks = {}

    # Check 1: Part number mentioned
    checks["part_mentioned"] = part_num.lower() in text.lower()

    # Check 2: Correct status
    checks["status_correct"] = expected_status.lower() in text.lower()

    # Check 3: Details handling
    if details_required and expected_details:
        checks["details_present"] = str(expected_details).lower() in text.lower()
    else:
        # Should NOT fabricate details
        fabrication_signals = [
            "additional details:",
            "details:",
        ]
        has_fabrication = any(sig in text.lower() for sig in fabrication_signals)
        if has_fabrication:
            # Check it's not just saying "no details" or similar
            no_detail_signals = ["no additional", "no details", "not available", "none"]
            actually_ok = any(ns in text.lower() for ns in no_detail_signals)
            checks["no_fabricated_details"] = actually_ok or not has_fabrication
        else:
            checks["no_fabricated_details"] = True

    all_passed = all(checks.values())

    return {
        "pass": all_passed,
        "expected": f"Part={part_num}, Status={expected_status}, Details={expected_details}",
        "got": text[:400],
        "criteria": f"Checks: {checks}",
        "strict_notes": "Verifies part number, exact status, and details handling.",
    }


def _eval_not_eligible(resp: dict, part_info: dict) -> dict:
    text = resp["text"]
    expected_status = part_info["status"]

    # Must mention the status
    status_present = expected_status.lower() in text.lower()

    # Must NOT independently say it CAN be scrapped (outside the status string itself).
    # Strip out the status string first, then check residual text for false positives.
    residual = text.lower().replace(expected_status.lower(), "")
    false_positive_phrases = [
        "can be scrapped",
        "eligible to be scrapped",
        "yes, it can",
        "you can scrap",
        "is eligible for scrap",
    ]
    falsely_says_yes = any(fp in residual for fp in false_positive_phrases)

    passed = status_present and not falsely_says_yes

    return {
        "pass": passed,
        "expected": f"NOT eligible: {expected_status}",
        "got": text[:400],
        "criteria": f"Status present={status_present}, false_positive_in_residual={falsely_says_yes}",
        "strict_notes": "Must show correct NOT-eligible status AND must not falsely say scrap is OK outside the status label.",
    }


def _eval_fake_part(resp: dict, fake_part: str) -> dict:
    text = resp["text"]
    not_found_phrases = [
        "could not find",
        "not found",
        "no data",
        "does not exist",
        "doesn't exist",
        "no matching",
        "no results",
        "cannot determine",
        "no record",
        "not in",
        "unavailable",
        "confirm if the part number is correct",
        "issue retrieving",
        "no rows",
    ]
    detected = any(p in text.lower() for p in not_found_phrases)

    # Also must NOT provide a fake status
    fabricated_status = any(
        s.lower() in text.lower()
        for s in [
            "May be eligible",
            "NOT eligible",
            "No stock",
            "In WhereUsed",
            "Open WorkOrder",
            "Open Sales Order",
            "Product USAGE",
            "Component Request",
            "Need Further Review",
            "Sold in Past",
            "REPAIR USAGE",
            "Custom Button",
            "International Powercord",
            "Bin Stock",
            "Bin Location",
        ]
    )

    passed = detected and not fabricated_status

    return {
        "pass": passed,
        "expected": "Part not found / no data",
        "got": text[:400],
        "criteria": f"Not-found detected={detected}, fabricated_status={fabricated_status}",
        "strict_notes": "Must say not found AND must not hallucinate a status.",
    }


def _eval_status_distribution(resp: dict, gt: dict) -> dict:
    text = resp["text"]
    data_chunks = resp.get("data_chunks", [])
    combined = text + " " + " ".join(data_chunks)

    # Check that the top 3 statuses appear somewhere
    top_statuses = sorted(gt["status_distribution"].items(), key=lambda x: -x[1])[:3]
    found_statuses = sum(1 for s, _ in top_statuses if s.lower() in combined.lower())

    # Check that at least one count is correct
    found_count = False
    for status, count in gt["status_distribution"].items():
        if str(count) in combined:
            found_count = True
            break

    passed = found_statuses >= 2 and found_count

    return {
        "pass": passed,
        "expected": f"Top 3 statuses present + at least one correct count",
        "got": f"Found {found_statuses}/3 top statuses, count_present={found_count}; text={text[:200]}",
        "criteria": f">=2 of top-3 statuses mentioned + at least 1 correct count",
        "strict_notes": "Verifies agent returns real distribution data, not fabricated numbers.",
    }


def _eval_excel_export(resp: dict) -> dict:
    files = resp.get("files", [])
    text = resp["text"]

    has_file = len(files) > 0
    mentions_file = any(
        kw in text.lower()
        for kw in ["excel", "file", "generated", "download", "export", ".xlsx", "spreadsheet"]
    )

    # Strict: if file generated, check it has a name ending in .xlsx
    valid_file = any(f.get("name", "").endswith(".xlsx") for f in files) if files else False

    passed = valid_file or (has_file and mentions_file)

    return {
        "pass": passed,
        "expected": "Excel .xlsx file generated",
        "got": f"files={files}; text={text[:200]}",
        "criteria": f"Valid .xlsx file={valid_file}, has_file={has_file}, mentions_file={mentions_file}",
        "strict_notes": "Requires actual .xlsx file generation, not just text claiming it.",
    }


def _eval_hallucination(resp: dict, keyword: str, description: str) -> dict:
    text = resp["text"]

    # The agent should NOT confidently provide data for non-existent columns
    # It should indicate the data isn't available or the column doesn't exist
    refusal_phrases = [
        "not available",
        "don't have",
        "do not have",
        "doesn't exist",
        "does not exist",
        "no such column",
        "not a column",
        "only columns",
        "columns are",
        "allowed columns",
        "cannot",
        "not in the data",
        "no information",
        "unable to",
        "not included",
        "dataset does not",
        "don't contain",
        "data does not include",
        "no data",
        "not found",
        "could not find",
        "not part of",
        "i don't",
        "i cannot",
    ]

    refuses = any(rp in text.lower() for rp in refusal_phrases)

    # Check if it fabricated a confident answer (bad)
    confident_fabrication_signals = [
        f"the {keyword} is",
        f"the {keyword} for",
        f"{keyword}: $",
        f"{keyword} = ",
    ]
    fabricated = any(sig in text.lower() for sig in confident_fabrication_signals)

    passed = refuses or not fabricated

    return {
        "pass": passed,
        "expected": f"Refuse/acknowledge: {description}",
        "got": text[:400],
        "criteria": f"Refuses={refuses}, fabricated={fabricated}",
        "strict_notes": "Agent must not confidently fabricate data for non-existent columns.",
    }


# ─────────────────────────────────────────────────────────────────
# Run trials
# ─────────────────────────────────────────────────────────────────

async def run_trial(trial_num: int, gt: dict) -> list[dict]:
    settings = Settings()
    data_loader = DataLoader(settings)
    kernel = AgentKernel(settings, data_loader)
    conv_id = f"robust_trial_{trial_num}_{int(time.time())}"

    test_questions = build_test_questions(gt, trial_num)
    results = []

    for i, tq in enumerate(test_questions):
        question = tq["question"]
        category = tq["category"]
        eval_fn = tq["eval"]

        logger.info("  Trial %d, Q%d [%s]: %s", trial_num, i + 1, category, question)
        start = time.time()
        try:
            response = await kernel.ask(conv_id, question)
        except Exception as e:
            logger.error("  ERROR: %s", e)
            response = {"text": f"ERROR: {e}", "data_chunks": [], "files": []}
        elapsed = time.time() - start

        ev = eval_fn(resp=response)
        ev["trial"] = trial_num
        ev["q_num"] = i + 1
        ev["category"] = category
        ev["question"] = question
        ev["latency_s"] = round(elapsed, 2)
        ev["raw_response"] = response.get("text", "")[:500]
        results.append(ev)

    return results


# ─────────────────────────────────────────────────────────────────
# Reports
# ─────────────────────────────────────────────────────────────────

def generate_markdown(all_results: list[dict], gt: dict, total_elapsed: float) -> str:
    df = pd.DataFrame(all_results)
    lines = []
    lines.append("# Robust Agent Benchmark Report (v2)")
    lines.append(f"\n**Date:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"**Trials:** {NUM_TRIALS}")
    lines.append(f"**Questions per trial:** {len(CATEGORIES)}")
    lines.append(f"**Total questions:** {len(df)}")
    lines.append(f"**Total runtime:** {total_elapsed:.1f}s")
    lines.append(f"**Model:** {Settings().azure_openai_deployment_name}")
    lines.append(f"**Table:** {Settings().sql_table} ({gt['total_rows']} rows)")

    # Overall accuracy
    total_pass = int(df["pass"].sum())
    total_q = len(df)
    overall = round(total_pass / total_q * 100, 1)
    lines.append(f"\n## Overall Accuracy: **{overall}%** ({total_pass}/{total_q})")

    # Per-category accuracy
    lines.append("\n## Per-Category Accuracy")
    lines.append("| Category | Pass Rate | Avg Latency | Notes |")
    lines.append("|----------|-----------|-------------|-------|")
    for cat in CATEGORIES:
        cdf = df[df["category"] == cat]
        if len(cdf) == 0:
            continue
        pr = cdf["pass"].sum() / len(cdf) * 100
        al = cdf["latency_s"].mean()
        lines.append(
            f"| {cat} | {pr:.0f}% ({int(cdf['pass'].sum())}/{len(cdf)}) | {al:.2f}s | |"
        )

    # Per-trial breakdown
    lines.append("\n## Per-Trial Breakdown")
    cats_short = ["Q1:Count", "Q2:Details", "Q3:NullDet", "Q4:NotElig",
                  "Q5:Fake", "Q6:Distrib", "Q7:Excel", "Q8:Halluc"]
    header = "| Trial | " + " | ".join(cats_short) + " | Score | Latency |"
    sep = "|-------" + "|------" * len(cats_short) + "|-------|---------|"
    lines.append(header)
    lines.append(sep)
    for trial in range(1, NUM_TRIALS + 1):
        tdf = df[df["trial"] == trial].sort_values("q_num")
        marks = []
        for _, row in tdf.iterrows():
            marks.append("✅" if row["pass"] else "❌")
        while len(marks) < len(cats_short):
            marks.append("—")
        score = int(tdf["pass"].sum())
        total_lat = tdf["latency_s"].sum()
        lines.append(f"| {trial} | {' | '.join(marks)} | {score}/{len(tdf)} | {total_lat:.1f}s |")

    # Detailed question log (all trials)
    lines.append("\n## Detailed Question Log")
    lines.append("| Trial | Q# | Category | Question | Pass | Latency |")
    lines.append("|-------|----|----------|----------|------|---------|")
    for _, row in df.iterrows():
        mark = "✅" if row["pass"] else "❌"
        lines.append(
            f"| {row['trial']} | {row['q_num']} | {row['category'][:25]} | "
            f"{row['question'][:50]} | {mark} | {row['latency_s']:.1f}s |"
        )

    # Failure details
    failures = df[~df["pass"]]
    if len(failures) > 0:
        lines.append(f"\n## Failure Details ({len(failures)} failures)")
        for _, row in failures.iterrows():
            lines.append(f"\n### Trial {row['trial']}, Q{row['q_num']}: {row['category']}")
            lines.append(f"- **Question:** {row['question']}")
            lines.append(f"- **Expected:** {row['expected']}")
            lines.append(f"- **Got:** {row['raw_response'][:400]}")
            lines.append(f"- **Criteria:** {row['criteria']}")
            if row.get("strict_notes"):
                lines.append(f"- **Strictness:** {row['strict_notes']}")
    else:
        lines.append("\n## No failures — all questions passed in all trials! 🎉")

    # Methodology
    lines.append("\n## Methodology")
    lines.append("- **Question variation:** Each trial uses different phrasings (10 variants per category)")
    lines.append("- **Part rotation:** Real parts rotate across trials (different statuses, with/without Details)")
    lines.append("- **Fake parts:** 10 different fake part numbers, verified not in DB")
    lines.append("- **Hallucination tests:** Asks about non-existent columns (price, supplier, warranty, etc.)")
    lines.append("- **Strict evaluation:** Part lookups verify exact status match + Details handling")
    lines.append("- **NOT-eligible test:** Checks agent doesn't falsely say a part can be scrapped")
    lines.append("- **Distribution test:** Verifies real counts appear, not fabricated numbers")
    lines.append("- **Excel test:** Requires actual .xlsx file generation, not just text claim")

    return "\n".join(lines)


def generate_excel(all_results: list[dict], gt: dict, output_path: str = None):
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    df = pd.DataFrame(all_results)
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filepath = output_path or os.path.join(OUTPUT_DIR, "benchmark_v2_results.xlsx")

    wb = Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    def style_header(ws, headers, row=1):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    def auto_width(ws, col_count):
        for c in range(1, col_count + 1):
            max_len = max(
                len(str(ws.cell(row=r, column=c).value or ""))
                for r in range(1, ws.max_row + 1)
            )
            ws.column_dimensions[get_column_letter(c)].width = min(max_len + 4, 55)

    # ── Sheet 1: Raw Results ──
    ws_raw = wb.active
    ws_raw.title = "Raw Results"
    raw_headers = ["Trial", "Q#", "Category", "Question", "Expected", "Got", "Pass", "Latency (s)", "Criteria"]
    style_header(ws_raw, raw_headers)

    for r_idx, (_, row) in enumerate(df.iterrows(), 2):
        vals = [
            row["trial"], row["q_num"], row["category"], row["question"][:80],
            str(row["expected"])[:80], row["raw_response"][:120],
            "PASS" if row["pass"] else "FAIL", row["latency_s"],
            str(row.get("criteria", ""))[:80],
        ]
        for c, v in enumerate(vals, 1):
            cell = ws_raw.cell(row=r_idx, column=c, value=v)
            cell.border = thin_border
            if c == 7:
                cell.fill = pass_fill if v == "PASS" else fail_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
    auto_width(ws_raw, len(raw_headers))

    # ── Sheet 2: Category Summary ──
    ws_cat = wb.create_sheet("Category Summary")
    ws_cat.cell(row=1, column=1, value="Per-Category Accuracy").font = Font(bold=True, size=14)
    cat_headers = ["Category", "Pass Count", "Fail Count", "Pass Rate (%)", "Avg Latency (s)"]
    style_header(ws_cat, cat_headers, row=3)

    for i, cat in enumerate(CATEGORIES):
        cdf = df[df["category"] == cat]
        if len(cdf) == 0:
            continue
        r = 4 + i
        pc = int(cdf["pass"].sum())
        fc = len(cdf) - pc
        pr = round(pc / len(cdf) * 100, 1)
        al = round(cdf["latency_s"].mean(), 2)
        vals = [cat, pc, fc, pr, al]
        for c, v in enumerate(vals, 1):
            cell = ws_cat.cell(row=r, column=c, value=v)
            cell.border = thin_border
            if c == 4:
                cell.fill = pass_fill if pr >= 80 else fail_fill
                cell.font = Font(bold=True)

    # Overall
    r_overall = 4 + len(CATEGORIES) + 1
    total_pass = int(df["pass"].sum())
    total_q = len(df)
    overall = round(total_pass / total_q * 100, 1)
    ws_cat.cell(row=r_overall, column=1, value="OVERALL").font = Font(bold=True, size=12)
    cell = ws_cat.cell(row=r_overall, column=2, value=f"{overall}% ({total_pass}/{total_q})")
    cell.font = Font(bold=True, size=12, color="4472C4")

    auto_width(ws_cat, len(cat_headers))

    # Chart: Pass rate by category
    chart1 = BarChart()
    chart1.type = "col"
    chart1.title = "Pass Rate by Category (%)"
    chart1.y_axis.title = "Pass Rate (%)"
    chart1.y_axis.scaling.min = 0
    chart1.y_axis.scaling.max = 100
    chart1.style = 10
    chart1.width = 24
    chart1.height = 14
    data_ref = Reference(ws_cat, min_col=4, min_row=3, max_row=3 + len(CATEGORIES))
    cats_ref = Reference(ws_cat, min_col=1, min_row=4, max_row=3 + len(CATEGORIES))
    chart1.add_data(data_ref, titles_from_data=True)
    chart1.set_categories(cats_ref)
    ws_cat.add_chart(chart1, f"A{r_overall + 3}")

    # ── Sheet 3: Per-Trial Scores ──
    ws_trial = wb.create_sheet("Per-Trial Scores")
    cats_short = ["Count", "Details", "NullDet", "NotElig", "Fake", "Distrib", "Excel", "Halluc"]
    trial_headers = ["Trial"] + cats_short + ["Score", "Total Latency (s)"]
    style_header(ws_trial, trial_headers)

    for trial in range(1, NUM_TRIALS + 1):
        tdf = df[df["trial"] == trial].sort_values("q_num")
        r = trial + 1
        ws_trial.cell(row=r, column=1, value=trial).border = thin_border
        for q_idx, (_, row) in enumerate(tdf.iterrows()):
            passed = bool(row["pass"])
            cell = ws_trial.cell(row=r, column=q_idx + 2, value="PASS" if passed else "FAIL")
            cell.fill = pass_fill if passed else fail_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        score = int(tdf["pass"].sum())
        ws_trial.cell(row=r, column=len(cats_short) + 2, value=score).border = thin_border
        ws_trial.cell(row=r, column=len(cats_short) + 3,
                       value=round(tdf["latency_s"].sum(), 1)).border = thin_border

    auto_width(ws_trial, len(trial_headers))

    # Chart: Score per trial
    chart2 = BarChart()
    chart2.type = "col"
    chart2.title = f"Score per Trial (out of {len(CATEGORIES)})"
    chart2.y_axis.title = "Questions Passed"
    chart2.y_axis.scaling.min = 0
    chart2.y_axis.scaling.max = len(CATEGORIES)
    chart2.style = 10
    chart2.width = 20
    chart2.height = 12
    score_ref = Reference(ws_trial, min_col=len(cats_short) + 2, min_row=1, max_row=NUM_TRIALS + 1)
    trial_cats = Reference(ws_trial, min_col=1, min_row=2, max_row=NUM_TRIALS + 1)
    chart2.add_data(score_ref, titles_from_data=True)
    chart2.set_categories(trial_cats)
    ws_trial.add_chart(chart2, "A14")

    # ── Sheet 4: Latency Analysis ──
    ws_lat = wb.create_sheet("Latency Analysis")
    ws_lat.cell(row=1, column=1, value="Avg Latency by Category").font = Font(bold=True, size=14)
    lat_headers = ["Category", "Min (s)", "Avg (s)", "Max (s)", "Std Dev (s)"]
    style_header(ws_lat, lat_headers, row=3)

    for i, cat in enumerate(CATEGORIES):
        cdf = df[df["category"] == cat]
        if len(cdf) == 0:
            continue
        r = 4 + i
        lats = cdf["latency_s"]
        vals = [cat, round(lats.min(), 2), round(lats.mean(), 2),
                round(lats.max(), 2), round(lats.std(), 2)]
        for c, v in enumerate(vals, 1):
            cell = ws_lat.cell(row=r, column=c, value=v)
            cell.border = thin_border
    auto_width(ws_lat, len(lat_headers))

    # Latency chart
    chart3 = BarChart()
    chart3.type = "col"
    chart3.title = "Average Latency by Category (seconds)"
    chart3.y_axis.title = "Seconds"
    chart3.style = 10
    chart3.width = 24
    chart3.height = 14
    lat_data = Reference(ws_lat, min_col=3, min_row=3, max_row=3 + len(CATEGORIES))
    lat_cats = Reference(ws_lat, min_col=1, min_row=4, max_row=3 + len(CATEGORIES))
    chart3.add_data(lat_data, titles_from_data=True)
    chart3.set_categories(lat_cats)
    ws_lat.add_chart(chart3, f"A{4 + len(CATEGORIES) + 2}")

    wb.save(filepath)
    return filepath


# ─────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────

async def main():
    logger.info("=" * 60)
    logger.info("ROBUST BENCHMARK v2: Fetching ground truth...")
    gt = fetch_ground_truth()
    logger.info("Ground truth: %d rows, %d statuses, %d parts with details",
                gt["total_rows"], len(gt["status_distribution"]),
                len(gt["parts_with_details"]))

    all_results = []
    overall_start = time.time()

    for trial in range(1, NUM_TRIALS + 1):
        logger.info("─" * 40)
        logger.info("TRIAL %d/%d", trial, NUM_TRIALS)
        trial_results = await run_trial(trial, gt)
        all_results.extend(trial_results)
        passed = sum(1 for r in trial_results if r["pass"])
        logger.info("  Trial %d result: %d/%d passed", trial, passed, len(trial_results))

    total_elapsed = time.time() - overall_start
    logger.info("=" * 60)
    logger.info("ALL TRIALS COMPLETE in %.1fs", total_elapsed)

    df = pd.DataFrame(all_results)
    total_pass = int(df["pass"].sum())
    total_q = len(df)
    logger.info("OVERALL ACCURACY: %d/%d (%.1f%%)", total_pass, total_q,
                total_pass / total_q * 100)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    md = generate_markdown(all_results, gt, total_elapsed)
    md_path = os.path.join(OUTPUT_DIR, "benchmark_v2_results.md")
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(md)
    logger.info("Markdown: %s", md_path)

    xlsx_path = os.path.join(OUTPUT_DIR, "benchmark_v2_results.xlsx")
    try:
        xlsx_path = generate_excel(all_results, gt)
    except PermissionError:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt = os.path.join(OUTPUT_DIR, f"benchmark_v2_results_{ts}.xlsx")
        xlsx_path = generate_excel(all_results, gt, output_path=alt)
    logger.info("Excel: %s", xlsx_path)

    print("\n" + "=" * 60)
    print(f"ROBUST BENCHMARK v2 COMPLETE — {total_pass}/{total_q} ({total_pass/total_q*100:.1f}%)")
    print(f"  Markdown: {md_path}")
    print(f"  Excel:    {xlsx_path}")
    print("=" * 60)


if __name__ == "__main__":
    asyncio.run(main())
